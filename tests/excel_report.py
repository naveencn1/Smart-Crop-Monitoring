import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

class ExcelReport:
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.excel_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'excel')
        os.makedirs(self.excel_dir, exist_ok=True)
        
        # Remove default sheet
        default_sheet = self.wb.active
        self.wb.remove(default_sheet)
        
        self.tests = []
        self.logs = []
        self.env_details = {"browser": "Chrome", "mode": "Headless", "url": "http://localhost:8000"}

    def set_env_details(self, browser, mode, url):
        self.env_details = {
            "browser": browser,
            "mode": mode,
            "url": url
        }

    def add_test(self, test_id, module, scenario, status, duration, start_time, end_time, failure_reason="", stack_trace="", screenshot_path=""):
        self.tests.append({
            "id": test_id,
            "module": module,
            "scenario": scenario,
            "status": status,
            "duration": duration, # in seconds
            "start_time": start_time,
            "end_time": end_time,
            "failure_reason": failure_reason,
            "stack_trace": stack_trace,
            "screenshot_path": screenshot_path
        })

    def add_log(self, timestamp, test_id, step, result, remarks=""):
        self.logs.append({
            "timestamp": timestamp,
            "test_id": test_id,
            "step": step,
            "result": result,
            "remarks": remarks
        })

    def generate_report(self):
        # Create worksheets
        summary_sheet = self.wb.create_sheet(title='Summary')
        test_cases_sheet = self.wb.create_sheet(title='Test Cases')
        failed_tests_sheet = self.wb.create_sheet(title='Failed Tests')
        execution_logs_sheet = self.wb.create_sheet(title='Execution Logs')
        
        # Styles
        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid') # Steel blue
        header_alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
        
        thin_border_side = Side(style='thin', color='E0E0E0')
        border_style = Border(top=thin_border_side, bottom=thin_border_side, left=thin_border_side, right=thin_border_side)
        
        # 1. Summary Sheet
        summary_sheet.views.sheetView[0].showGridLines = True
        summary_sheet.append([
            'Execution Date', 'Browser / Driver', 'Execution Mode', 'Target URL', 
            'Total Tests', 'Passed', 'Failed', 'Pass Percentage', 'Execution Duration'
        ])
        
        total = len(self.tests)
        passed = sum(1 for t in self.tests if t['status'].lower() == 'passed')
        failed = sum(1 for t in self.tests if t['status'].lower() == 'failed')
        pass_percentage = f"{round((passed / total) * 100)}%" if total > 0 else "0%"
        total_duration = sum(t['duration'] for t in self.tests)
        duration_formatted = f"{total_duration:.2f}s"
        
        summary_sheet.append([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            self.env_details['browser'],
            self.env_details['mode'],
            self.env_details['url'],
            total,
            passed,
            failed,
            pass_percentage,
            duration_formatted
        ])
        
        # Style Summary
        summary_sheet.row_dimensions[1].height = 28
        summary_sheet.row_dimensions[2].height = 24
        for col_idx in range(1, 10):
            cell = summary_sheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
            val_cell = summary_sheet.cell(row=2, column=col_idx)
            val_cell.font = Font(name='Segoe UI', size=10)
            val_cell.border = border_style
            val_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if col_idx == 8: # Pass Percentage coloring
                val_cell.font = Font(name='Segoe UI', size=10, bold=True, color='2E7D32' if passed == total and total > 0 else 'C62828')
                
        # 2. Test Cases Sheet
        test_cases_sheet.views.sheetView[0].showGridLines = True
        test_cases_sheet.append([
            'Test ID', 'Module', 'Scenario', 'Status', 'Start Time', 'End Time', 'Duration'
        ])
        test_cases_sheet.row_dimensions[1].height = 28
        
        for col_idx in range(1, 8):
            cell = test_cases_sheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
        for idx, t in enumerate(self.tests, start=2):
            test_cases_sheet.append([
                t['id'],
                t['module'],
                t['scenario'],
                t['status'],
                t['start_time'].strftime("%Y-%m-%d %H:%M:%S") if t['start_time'] else "",
                t['end_time'].strftime("%Y-%m-%d %H:%M:%S") if t['end_time'] else "",
                f"{t['duration']:.2f}s"
            ])
            test_cases_sheet.row_dimensions[idx].height = 22
            for col_idx in range(1, 8):
                cell = test_cases_sheet.cell(row=idx, column=col_idx)
                cell.font = Font(name='Segoe UI', size=10)
                cell.border = border_style
                cell.alignment = Alignment(vertical='center')
                if col_idx in [1, 4, 7]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                if col_idx == 4:
                    if t['status'].lower() == 'passed':
                        cell.font = Font(name='Segoe UI', bold=True, color='2E7D32')
                    elif t['status'].lower() == 'failed':
                        cell.font = Font(name='Segoe UI', bold=True, color='C62828')
                    else:
                        cell.font = Font(name='Segoe UI', bold=True, color='F9A825')

        # 3. Failed Tests Sheet
        failed_tests_sheet.views.sheetView[0].showGridLines = True
        failed_tests_sheet.append([
            'Test ID', 'Scenario', 'Failure Reason', 'Screenshot Path', 'Failure Details / Stack Trace'
        ])
        failed_tests_sheet.row_dimensions[1].height = 28
        
        failed_fill = PatternFill(start_color='C62828', end_color='C62828', fill_type='solid') # Red header
        for col_idx in range(1, 6):
            cell = failed_tests_sheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = failed_fill
            cell.alignment = header_alignment
            
        failed_list = [t for t in self.tests if t['status'].lower() == 'failed']
        for idx, t in enumerate(failed_list, start=2):
            failed_tests_sheet.append([
                t['id'],
                t['scenario'],
                t['failure_reason'],
                t['screenshot_path'],
                t['stack_trace']
            ])
            failed_tests_sheet.row_dimensions[idx].height = 26
            for col_idx in range(1, 6):
                cell = failed_tests_sheet.cell(row=idx, column=col_idx)
                cell.font = Font(name='Segoe UI', size=9)
                cell.border = border_style
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        # 4. Execution Logs Sheet
        execution_logs_sheet.views.sheetView[0].showGridLines = True
        execution_logs_sheet.append([
            'Timestamp', 'Test ID', 'Step Description', 'Result', 'Remarks'
        ])
        execution_logs_sheet.row_dimensions[1].height = 28
        for col_idx in range(1, 6):
            cell = execution_logs_sheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
        for idx, l in enumerate(self.logs, start=2):
            execution_logs_sheet.append([
                l['timestamp'],
                l['test_id'],
                l['step'],
                l['result'],
                l['remarks']
            ])
            execution_logs_sheet.row_dimensions[idx].height = 20
            for col_idx in range(1, 6):
                cell = execution_logs_sheet.cell(row=idx, column=col_idx)
                cell.font = Font(name='Segoe UI', size=9)
                cell.border = border_style
                cell.alignment = Alignment(vertical='center')
                if col_idx in [1, 2, 4]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                if col_idx == 4:
                    res = l['result'].lower()
                    if 'pass' in res or 'success' in res:
                        cell.font = Font(name='Segoe UI', size=9, bold=True, color='2E7D32')
                    elif 'fail' in res or 'error' in res:
                        cell.font = Font(name='Segoe UI', size=9, bold=True, color='C62828')
                        
        # Auto-adjust column widths for all sheets
        for sheet in [summary_sheet, test_cases_sheet, failed_tests_sheet, execution_logs_sheet]:
            for col in sheet.columns:
                max_len = 0
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                for cell in col:
                    if cell.value is not None:
                        # Handle multiple lines in cell for length estimation
                        lines = str(cell.value).split('\n')
                        for line in lines:
                            max_len = max(max_len, len(line))
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
        # Write file
        file_path = os.path.join(self.excel_dir, 'Web_E2E_Report.xlsx')
        self.wb.save(file_path)
        return file_path

excel_report = ExcelReport()

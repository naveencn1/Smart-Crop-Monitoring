import os
import time
import random
import requests
import openpyxl
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

BACKEND_URL = "http://localhost:5000"
TEST_DURATION = 60  # seconds
CONCURRENT_USERS = 100
EXCEL_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'excel')
REPORT_PATH = os.path.join(EXCEL_DIR, 'Load_Test_Report.xlsx')

# Target Endpoints
ENDPOINTS = [
    "/api/health",
    "/api/sensors",
    "/api/profile",
    "/api/alerts"
]

def send_request(session, endpoint):
    url = f"{BACKEND_URL}{endpoint}"
    start = time.perf_counter()
    try:
        response = session.get(url, timeout=3)
        duration = (time.perf_counter() - start) * 1000 # in ms
        status = "SUCCESS" if response.status_code in [200, 201] else "FAILED"
        return {
            "timestamp": datetime.now().strftime("%H:%M:%S.%f")[:-3],
            "endpoint": endpoint,
            "duration": duration,
            "status": status,
            "error": f"HTTP {response.status_code}" if status == "FAILED" else ""
        }
    except Exception as e:
        duration = (time.perf_counter() - start) * 1000
        return {
            "timestamp": datetime.now().strftime("%H:%M:%S.%f")[:-3],
            "endpoint": endpoint,
            "duration": duration,
            "status": "FAILED",
            "error": str(e)
        }

def run_user_session(stop_time):
    session = requests.Session()
    local_results = []
    
    while time.time() < stop_time:
        endpoint = random.choice(ENDPOINTS)
        res = send_request(session, endpoint)
        local_results.append(res)
        # Tiny delay to simulate real user click gap (e.g. 50ms)
        time.sleep(0.05)
        
    return local_results

def generate_excel_report(results, run_start, run_end, total_duration):
    os.makedirs(EXCEL_DIR, exist_ok=True)
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # 1. Calculation of statistics
    total_requests = len(results)
    successful_requests = sum(1 for r in results if r['status'] == 'SUCCESS')
    failed_requests = total_requests - successful_requests
    success_rate = f"{round((successful_requests / total_requests) * 100, 2)}%" if total_requests > 0 else "0%"
    
    durations = [r['duration'] for r in results]
    min_time = min(durations) if durations else 0
    max_time = max(durations) if durations else 0
    avg_time = sum(durations) / len(durations) if durations else 0
    
    rps = total_requests / total_duration if total_duration > 0 else 0
    
    # Setup Sheets
    summary_sheet = wb.create_sheet(title='Summary')
    details_sheet = wb.create_sheet(title='Requests Log')
    
    # Styles
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid') # Steel blue
    header_alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
    
    thin_border_side = Side(style='thin', color='E0E0E0')
    border_style = Border(top=thin_border_side, bottom=thin_border_side, left=thin_border_side, right=thin_border_side)
    
    # Determine if running in GitHub Actions to link to the workflow run
    github_run_url = "N/A"
    if os.environ.get("GITHUB_ACTIONS") == "true" or "GITHUB_RUN_ID" in os.environ:
        server = os.environ.get("GITHUB_SERVER_URL", "https://github.com")
        repo = os.environ.get("GITHUB_REPOSITORY", "")
        run_id = os.environ.get("GITHUB_RUN_ID", "")
        github_run_url = f"{server}/{repo}/actions/runs/{run_id}"

    # --- Summary Sheet ---
    summary_sheet.views.sheetView[0].showGridLines = True
    summary_headers = [
        'Execution Date', 'Target Host', 'Concurrent VUs', 'Duration (s)', 
        'Total Requests', 'Passed (Success)', 'Failed (Errors)', 'Success Rate', 
        'Avg RPS', 'Min Latency (ms)', 'Avg Latency (ms)', 'Max Latency (ms)',
        'GitHub Actions Run Link'
    ]
    summary_sheet.append(summary_headers)
    summary_sheet.append([
        run_start.strftime("%Y-%m-%d %H:%M:%S"),
        BACKEND_URL,
        CONCURRENT_USERS,
        f"{total_duration:.1f}",
        total_requests,
        successful_requests,
        failed_requests,
        success_rate,
        round(rps, 2),
        round(min_time, 1),
        round(avg_time, 1),
        round(max_time, 1),
        github_run_url
    ])
    
    # Format Summary Sheet
    summary_sheet.row_dimensions[1].height = 28
    summary_sheet.row_dimensions[2].height = 24
    for col_idx in range(1, len(summary_headers) + 1):
        cell = summary_sheet.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        
        val_cell = summary_sheet.cell(row=2, column=col_idx)
        val_cell.font = Font(name='Segoe UI', size=10)
        val_cell.border = border_style
        val_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        if col_idx == 8: # Success rate coloring
            val_cell.font = Font(name='Segoe UI', size=10, bold=True, color='2E7D32' if successful_requests == total_requests else 'C62828')
            
        # Link styling for GitHub Actions Run URL
        if col_idx == 13 and github_run_url != "N/A":
            val_cell.font = Font(name='Segoe UI', size=10, color='0563C1', underline='single')
            
    # --- Details Sheet ---
    details_sheet.views.sheetView[0].showGridLines = True
    details_headers = ['Timestamp', 'Endpoint', 'Response Time (ms)', 'Status', 'Error Details']
    details_sheet.append(details_headers)
    details_sheet.row_dimensions[1].height = 28
    
    for col_idx in range(1, len(details_headers) + 1):
        cell = details_sheet.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        
    # Cap detailed requests logged to 1000 rows to keep excel readable and small
    logged_results = results[:1000]
    for idx, r in enumerate(logged_results, start=2):
        details_sheet.append([
            r['timestamp'],
            r['endpoint'],
            round(r['duration'], 1),
            r['status'],
            r['error']
        ])
        details_sheet.row_dimensions[idx].height = 20
        for col_idx in range(1, len(details_headers) + 1):
            cell = details_sheet.cell(row=idx, column=col_idx)
            cell.font = Font(name='Segoe UI', size=9)
            cell.border = border_style
            cell.alignment = Alignment(vertical='center')
            if col_idx in [1, 4]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            if col_idx == 4:
                if r['status'] == 'SUCCESS':
                    cell.font = Font(name='Segoe UI', size=9, bold=True, color='2E7D32')
                else:
                    cell.font = Font(name='Segoe UI', size=9, bold=True, color='C62828')
                    
    # Auto-adjust column widths
    for sheet in [summary_sheet, details_sheet]:
        for col in sheet.columns:
            max_len = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    wb.save(REPORT_PATH)
    print(f"[Load Test] Excel report saved to {REPORT_PATH}")

def main():
    print(f"=== Commencing API Load Testing ===")
    print(f"  Target: {BACKEND_URL}")
    print(f"  Concurrent Virtual Users: {CONCURRENT_USERS}")
    print(f"  Execution Duration: {TEST_DURATION}s")
    
    # Warm up ping
    try:
        requests.get(f"{BACKEND_URL}/api/health", timeout=2)
    except Exception:
        print(f"[-] Error: Backend server is not running on {BACKEND_URL}. Start the Flask app first.")
        return
        
    run_start = datetime.now()
    start_time = time.time()
    stop_time = start_time + TEST_DURATION
    
    results = []
    with ThreadPoolExecutor(max_workers=CONCURRENT_USERS) as executor:
        futures = [executor.submit(run_user_session, stop_time) for _ in range(CONCURRENT_USERS)]
        for fut in as_completed(futures):
            results.extend(fut.result())
            
    run_end = datetime.now()
    total_duration = time.time() - start_time
    
    print("\n=== Load Test Execution Finished ===")
    print(f"  Total Requests Sent: {len(results)}")
    print(f"  Failed Requests: {sum(1 for r in results if r['status'] == 'FAILED')}")
    print(f"  RPS (Requests/Sec): {round(len(results) / total_duration, 2)}")
    
    if os.environ.get("GITHUB_ACTIONS") == "true" or "GITHUB_RUN_ID" in os.environ:
        server = os.environ.get("GITHUB_SERVER_URL", "https://github.com")
        repo = os.environ.get("GITHUB_REPOSITORY", "")
        run_id = os.environ.get("GITHUB_RUN_ID", "")
        print(f"  GitHub Actions Run: {server}/{repo}/actions/runs/{run_id}")
        
    # Compile statistics and generate Excel report
    generate_excel_report(results, run_start, run_end, total_duration)

if __name__ == "__main__":
    main()

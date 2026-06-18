import os
import json
import subprocess
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

def run_bandit():
    json_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'bandit_raw.json')
    # Run bandit scan
    cmd = [
        "bandit",
        "-r",
        "backend/",
        "tests/",
        "-ll",
        "-f",
        "json",
        "-o",
        json_path
    ]
    # Run bandit command
    result = subprocess.run(cmd, capture_output=True, text=True)
    return json_path, result.returncode

def generate_excel_report(json_path):
    # Load json data
    if not os.path.exists(json_path):
        raise FileNotFoundError(f"Bandit scan output file not found: {json_path}")
        
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
        
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Create sheets
    summary_sheet = wb.create_sheet(title='Summary')
    findings_sheet = wb.create_sheet(title='Security Findings')
    
    # Styling
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid') # Deep blue
    header_alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
    
    thin_border_side = Side(style='thin', color='E0E0E0')
    border_style = Border(top=thin_border_side, bottom=thin_border_side, left=thin_border_side, right=thin_border_side)
    
    # 1. Summary sheet
    summary_sheet.views.sheetView[0].showGridLines = True
    summary_sheet.append([
        'Scan Date', 'Tool Used', 'Target Folders', 
        'Total Issues', 'High Severity', 'Medium Severity', 'Low Severity'
    ])
    
    total_issues = len(data.get('results', []))
    
    # Count severities in results
    high_count = sum(1 for r in data.get('results', []) if r.get('issue_severity') == 'HIGH')
    medium_count = sum(1 for r in data.get('results', []) if r.get('issue_severity') == 'MEDIUM')
    low_count = sum(1 for r in data.get('results', []) if r.get('issue_severity') == 'LOW')
    
    summary_sheet.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Bandit (Python SAST)",
        "backend/, tests/",
        total_issues,
        high_count,
        medium_count,
        low_count
    ])
    
    summary_sheet.row_dimensions[1].height = 28
    summary_sheet.row_dimensions[2].height = 24
    for col_idx in range(1, 8):
        cell = summary_sheet.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        
        val_cell = summary_sheet.cell(row=2, column=col_idx)
        val_cell.font = Font(name='Segoe UI', size=10)
        val_cell.border = border_style
        val_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Color total issues cell based on presence of issues
        if col_idx == 4:
            if total_issues > 0:
                val_cell.font = Font(name='Segoe UI', size=10, bold=True, color='C62828')
            else:
                val_cell.font = Font(name='Segoe UI', size=10, bold=True, color='2E7D32')
                
    # 2. Findings sheet (Note: Absolutely no filename or path columns)
    findings_sheet.views.sheetView[0].showGridLines = True
    findings_sheet.append([
        'Issue ID', 'Line Number', 'Severity', 'Confidence', 'Description', 'CWE Reference'
    ])
    findings_sheet.row_dimensions[1].height = 28
    
    for col_idx in range(1, 7):
        cell = findings_sheet.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        
    for idx, r in enumerate(data.get('results', []), start=2):
        findings_sheet.append([
            r.get('test_id'),
            r.get('line_number'),
            r.get('issue_severity'),
            r.get('issue_confidence'),
            r.get('issue_text'),
            r.get('ref')
        ])
        findings_sheet.row_dimensions[idx].height = 24
        
        for col_idx in range(1, 7):
            cell = findings_sheet.cell(row=idx, column=col_idx)
            cell.font = Font(name='Segoe UI', size=10)
            cell.border = border_style
            cell.alignment = Alignment(vertical='center')
            
            if col_idx in [1, 2, 3, 4]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
            # Color severity
            if col_idx == 3:
                sev = str(cell.value).upper()
                if sev == 'HIGH':
                    cell.font = Font(name='Segoe UI', size=10, bold=True, color='C62828')
                elif sev == 'MEDIUM':
                    cell.font = Font(name='Segoe UI', size=10, bold=True, color='EF6C00')
                else:
                    cell.font = Font(name='Segoe UI', size=10, bold=True, color='F9A825')
                    
    # Auto-fit column widths
    for sheet in [summary_sheet, findings_sheet]:
        for col in sheet.columns:
            max_len = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    lines = str(cell.value).split('\n')
                    for line in lines:
                        max_len = max(max_len, len(line))
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    # Save Excel report
    excel_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'excel')
    os.makedirs(excel_dir, exist_ok=True)
    report_path = os.path.join(excel_dir, 'Security_Scan_Report.xlsx')
    wb.save(report_path)
    
    # Clean up raw JSON file
    if os.path.exists(json_path):
        os.remove(json_path)
        
    return report_path

def print_qr_and_links(title, url):
    import sys
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass
    
    print("\n" + "=" * 50)
    print(f" {title.upper()} ".center(50, "="))
    print(f"URL Link: {url}")
    print("=" * 50)
    print("Scan QR Code to open:")
    try:
        import qrcode
        qr = qrcode.QRCode(version=1, box_size=1, border=1)
        qr.add_data(url)
        qr.make(fit=True)
        qr.print_ascii(out=sys.stdout, invert=True)
    except Exception as e:
        print(f"[Notice] Could not print QR Code: {e}")
    print("=" * 50 + "\n")

if __name__ == '__main__':
    print("[SAST Security Scan] Running static analysis...")
    json_path, code = run_bandit()
    
    try:
        report_path = generate_excel_report(json_path)
        print("[SAST Security Scan] Excel report generated successfully.")
    except Exception as e:
        print(f"[SAST Security Scan] Error generating Excel report: {e}")
        if os.path.exists(json_path):
            os.remove(json_path)
            
    print_qr_and_links("GitHub Actions Workflow", "https://github.com/naveencn1/Smart-Crop-Monitoring/actions")
    exit(code)

